from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app  # Add current_app here
from flask_login import login_required, current_user
from models import db, Part, Category, Supplier, BinCard, Warehouse, WarehouseStock, PartName
from werkzeug.utils import secure_filename
from views.utils import role_required
import os
from sqlalchemy.exc import IntegrityError
from sqlalchemy import desc
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
from flask import send_file
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from forms import PartForm  # Updated import statement
from utils.part_code_generator import generate_part_code


parts = Blueprint('parts', __name__)

# Configure allowed file extensions
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def is_user_staff():
    """Check if current user has staff role"""
    try:
        # First try checking direct role attribute
        if hasattr(current_user, 'role'):
            return current_user.role == 'staff'
        # Then try checking roles relationship
        elif hasattr(current_user, 'roles'):
            return any(role.name == 'staff' for role in current_user.roles)
        # Finally try checking role_id
        elif hasattr(current_user, 'role_id'):
            return current_user.role_id == 2  # Assuming 2 is staff role_id
        return False
    except:
        return False

@parts.route('/parts/staff')
@login_required
def staff_list_parts():
    page = request.args.get('page', 1, type=int)
    # Get all parts ordered by creation date (newest first)
    pagination = Part.query.order_by(desc(Part.created_at)).paginate(page=page, per_page=10, error_out=False)
    parts = pagination.items
    
    # Update stock levels from all warehouses
    for part in parts:
        # Calculate total stock across all warehouses
        total_stock = sum(stock.quantity for stock in part.warehouse_stocks)
        part.stock_level = total_stock
    
    # Get categories for filtering
    categories = Category.query.all()
    
    return render_template('parts/staff_list.html', parts=parts, categories=categories, pagination=pagination)

@parts.route('/parts')
@login_required
def list_parts():
    page = request.args.get('page', 1, type=int)
    # Get all parts with their warehouse stocks
    q = request.args.get('q', '').strip()
    category_id = request.args.get('category_id')
    query = Part.query
    if q:
        # Join with related tables for searching
        # Using outerjoin to include parts that might not have these relationships
        query = query.outerjoin(Category, Part.category_id == Category.id)\
                     .outerjoin(Supplier, Part.supplier_id == Supplier.id)\
                     .outerjoin(WarehouseStock, Part.id == WarehouseStock.part_id)\
                     .outerjoin(Warehouse, WarehouseStock.warehouse_id == Warehouse.id)

        query = query.filter(
            (Part.part_number.ilike(f'%{q}%')) |
            (Part.name.ilike(f'%{q}%')) |
            (Part.location.ilike(f'%{q}%')) |
            (Part.code.ilike(f'%{q}%')) |
            (Part.substitute_part_number.ilike(f'%{q}%')) |
            (Part.description.ilike(f'%{q}%')) |
            (Part.weight.cast(db.String).ilike(f'%{q}%')) |
            (Part.unit.ilike(f'%{q}%')) |
            (Part.barcode.ilike(f'%{q}%')) |
            (Part.created_at.cast(db.String).ilike(f'%{q}%')) |
            (Part.updated_at.cast(db.String).ilike(f'%{q}%')) |
            (Category.name.ilike(f'%{q}%')) |
            (Supplier.name.ilike(f'%{q}%')) |
            (Warehouse.name.ilike(f'%{q}%'))
        )
    if category_id:
        query = query.filter(Part.category_id == category_id)

    # Use distinct to avoid duplicate results from joins
    pagination = query.distinct().order_by(Part.created_at.desc()).paginate(page=page, per_page=10, error_out=False)
    parts = pagination.items
    
    # Update stock levels from all warehouses
    for part in parts:
        # Calculate total stock across all warehouses
        total_stock = sum(stock.quantity for stock in part.warehouse_stocks)
        
        # Update the part's stock level
        part.stock_level = total_stock
        
        # Get the latest bincard entry
        latest_bincard = BinCard.query.filter_by(part_id=part.id)\
            .order_by(BinCard.date.desc())\
            .first()
        
        # If there's a bincard entry and its balance doesn't match total stock,
        # create a new adjustment entry
        if latest_bincard and latest_bincard.balance != total_stock:
            adjustment = BinCard(
                part_id=part.id,
                transaction_type='adjustment',
                quantity=abs(total_stock - latest_bincard.balance),
                reference_type='stock_reconciliation',
                reference_id=0,
                balance=total_stock,
                user_id=current_user.id,
                notes=f'Stock reconciliation: adjusted to match warehouse totals'
            )
            db.session.add(adjustment)
    
    # Commit the changes to update stock levels and bincard entries
    db.session.commit()
    
    categories = Category.query.all()
    return render_template('parts/list.html', parts=parts, categories=categories, pagination=pagination)

def generate_sequential_barcode():
    # Get the highest existing barcode from the database
    last_part = Part.query.order_by(Part.barcode.desc()).first()
    if last_part and last_part.barcode.isdigit():
        # Increment the highest barcode by 1
        return str(int(last_part.barcode) + 1).zfill(10)  # Ensure it's 10 digits
    else:
        # Start the sequence if no barcodes exist
        return '1000000000'  # Start from 1000000000


@parts.route('/parts/add', methods=['GET', 'POST'])
@login_required
def add_part():
    # Get current user's staff status
    is_staff = is_user_staff()
    print(f"User {current_user.username} is staff: {is_staff}")  # Debug line
    
    form = PartForm()
    print(f"Form data: {request.form}")  # Debug line
    print(f"Form errors: {form.errors}")  # Debug line
    
    if request.method == 'POST':
        print("POST request received")  # Debug line
        print(f"Form data: {request.form}")  # Debug line
        print(f"Files: {request.files}")  # Debug line
        
        if form.validate_on_submit():
            try:
                print("Form validated successfully")  # Debug line
                # Generate part code
                part_code = generate_part_code(
                    form.manufacturer.data,
                    form.quality_level.data,
                    form.part_number.data
                )
                print(f"Generated part code: {part_code}")  # Debug line

                # Handle barcode properly
                barcode = request.form.get('barcode', '').strip()
                if barcode:  # Use the provided barcode if it's not empty
                    # Validate barcode uniqueness
                    existing_part = Part.query.filter_by(barcode=barcode).first()
                    if existing_part:
                        flash('Barcode must be unique. Another part already uses this barcode.', 'error')
                        return redirect(url_for('parts.add_part'))
                else:
                    # Generate a sequential barcode if none is provided
                    barcode = generate_sequential_barcode()

                # Create new part
                part = Part(
                    part_number=form.part_number.data,
                    name=form.name.data,
                    manufacturer=form.manufacturer.data,
                    quality_level=form.quality_level.data,
                    location=form.location.data,
                    code=part_code,  # Use the generated part code
                    substitute_part_number=request.form.get('substitute_part_number'),
                    stock_level=int(request.form.get('initial_stock', 0)),
                    min_stock=int(request.form.get('min_stock', 0)),
                    weight=float(request.form.get('weight')) if request.form.get('weight') else None,
                    height=float(request.form.get('height')) if request.form.get('height') else None,
                    length=float(request.form.get('length')) if request.form.get('length') else None,
                    width=float(request.form.get('width')) if request.form.get('width') else None,
                    color=request.form.get('color'),
                    description=form.description.data,
                    unit=request.form.get('unit'),
                    category_id=request.form.get('category_id') or None,
                    supplier_id=request.form.get('supplier_id') or None,
                    barcode=barcode,
                    cost_price=form.cost_price.data,
                    cost_price_dirham=form.cost_price_dirham.data,  
                    selling_price=form.selling_price.data
                )
                print(f"Created part object: {part}")  # Debug line
                
                # Only set price fields if user is not staff
                if not is_staff:
                    part.min_price = float(request.form.get('min_price')) if request.form.get('min_price') else None
                    part.max_price = float(request.form.get('max_price')) if request.form.get('max_price') else None
                
                # Handle image upload
                if 'image' in request.files:
                    file = request.files['image']
                    if file and allowed_file(file.filename):
                        try:
                            filename = secure_filename(file.filename)
                            upload_folder = os.path.join(current_app.root_path, 'static', 'uploads')
                            
                            # Create upload folder if it doesn't exist
                            if not os.path.exists(upload_folder):
                                os.makedirs(upload_folder)
                                
                            file.save(os.path.join(upload_folder, filename))
                            part.image_url = f'/static/uploads/{filename}'
                        except Exception as e:
                            flash(f'Error uploading image: {str(e)}', 'error')
                            return redirect(url_for('parts.add_part'))
                
                # Add part to database first
                db.session.add(part)
                db.session.flush()  # Flush to get the part ID
                print(f"Part added to session, ID: {part.id}")  # Debug line
                
                # Add the part name to PartName table if not already present
                name = form.name.data.strip()
                if not PartName.query.filter_by(name=name).first():
                    db.session.add(PartName(name=name))
                
                # Create warehouse stock entry
                warehouse_id = request.form.get('warehouse_id')
                if not warehouse_id:
                    flash('Warehouse selection is required', 'error')
                    db.session.rollback()
                    return redirect(url_for('parts.add_part'))
                    
                warehouse = Warehouse.query.get(warehouse_id)
                if not warehouse:
                    flash('Selected warehouse not found', 'error')
                    db.session.rollback()
                    return redirect(url_for('parts.add_part'))
                
                warehouse_stock = WarehouseStock(
                    warehouse_id=warehouse_id,
                    part_id=part.id,
                    quantity=int(request.form.get('initial_stock', 0))
                )
                db.session.add(warehouse_stock)
                db.session.flush()
                
                # Create bin card entry for initial stock
                initial_stock = int(request.form.get('initial_stock', 0))
                if initial_stock > 0:
                    bin_card = BinCard(
                        part_id=part.id,
                        transaction_type='in',
                        quantity=initial_stock,
                        reference_type='initial_stock',
                        reference_id=warehouse_stock.id,
                        balance=initial_stock,
                        user_id=current_user.id,
                        notes=f'Initial stock in warehouse {warehouse.name}'
                    )
                    db.session.add(bin_card)
                
                db.session.commit()
                print("Database commit successful")  # Debug line
                flash('Part added successfully', 'success')
                return redirect(url_for('parts.list_parts'))
                
            except IntegrityError as e:
                print(f"IntegrityError: {str(e)}")  # Debug line
                db.session.rollback()
                flash(f'Database error: {str(e)}', 'error')
                return redirect(url_for('parts.add_part'))
            except Exception as e:
                print(f"Exception: {str(e)}")  # Debug line
                db.session.rollback()
                flash(f'Error adding part: {str(e)}', 'error')
                return redirect(url_for('parts.add_part'))
        else:
            print("Form validation failed")  # Debug line
            print(f"Form errors: {form.errors}")  # Debug line
            flash('Please correct the errors in the form', 'error')
            
    categories = Category.query.all()
    suppliers = Supplier.query.all()
    warehouses = Warehouse.query.all()
    return render_template('parts/add.html', 
                         form=form,
                         categories=categories, 
                         suppliers=suppliers, 
                         warehouses=warehouses,
                         is_staff=is_staff)

@parts.route('/parts/<int:part_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_part(part_id):
    part = Part.query.get_or_404(part_id)
    is_staff = is_user_staff()
    print(f"User {current_user.username} is staff: {is_staff}")  # Debug line
    
    if request.method == 'POST':
        try:
            # Debug: Print form data
            print(request.form)
            

            # Update non-price fields
            part.name = request.form.get('name')
            part.substitute_part_number = request.form.get('substitute_part_number')
            part.code = request.form.get('code')
            part.location = request.form.get('location')
            part.description = request.form.get('description')
            part.unit = request.form.get('unit')
            part.min_stock = int(request.form.get('min_stock', 0))
            part.color = request.form.get('color', '')
            part.weight = float(request.form.get('weight', 0) or 0) if request.form.get('weight') else None
            part.height = float(request.form.get('height', 0) or 0) if request.form.get('height') else None
            part.length = float(request.form.get('length', 0) or 0) if request.form.get('length') else None
            part.width = float(request.form.get('width', 0) or 0) if request.form.get('width') else None
            part.warranty_period = int(request.form.get('warranty_period', 0) or 0)  # Handle empty string
            part.warranty_period = int(request.form.get('warranty_period', 0) or 0)  # Handle empty string
            
           # Handle barcode properly
            barcode = request.form.get('barcode', '').strip()
            if barcode:  # Only update if a new barcode is provided
                # Validate barcode uniqueness
                existing_part = Part.query.filter(Part.barcode == barcode, Part.id != part_id).first()
                if existing_part:
                    flash('Barcode must be unique. Another part already uses this barcode.', 'error')
                    return redirect(url_for('parts.edit_part', part_id=part_id))
                part.barcode = barcode
            else:
                part.barcode = None  # Set to NULL if no barcode is provided

            # Only update price fields if not staff
            if not is_staff:
                if request.form.get('min_price'):
                    part.min_price = float(request.form.get('min_price'))
                if request.form.get('max_price'):
                    part.max_price = float(request.form.get('max_price'))
                if request.form.get('cost_price'):
                    part.cost_price = float(request.form.get('cost_price'))
                if request.form.get('cost_price_dirham'):
                    part.cost_price_dirham = float(request.form.get('cost_price_dirham'))
                if request.form.get('selling_price'):
                    part.selling_price = float(request.form.get('selling_price'))
            
             # Update warehouse stock quantities
            for stock in part.warehouse_stocks:
                stock_quantity = request.form.get(f'stock_{stock.warehouse_id}')
                if stock_quantity is not None:
                    stock.quantity = int(stock_quantity)

            # Handle new warehouse stock
            new_warehouse_id = request.form.get('new_warehouse_id')
            new_warehouse_quantity = request.form.get('new_warehouse_quantity')
            if new_warehouse_id and new_warehouse_quantity:
                new_stock = WarehouseStock(
                    warehouse_id=new_warehouse_id,
                    part_id=part.id,
                    quantity=int(new_warehouse_quantity)
                )
                db.session.add(new_stock)

            # Handle image upload
            if 'image' in request.files:
                file = request.files['image']
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    upload_folder = os.path.join(current_app.root_path, 'static', 'uploads')
                    if not os.path.exists(upload_folder):
                        os.makedirs(upload_folder)
                    file.save(os.path.join(upload_folder, filename))
                    part.image_url = f'/static/uploads/{filename}'

            # Commit changes to the database
            with db.session.no_autoflush:
                db.session.commit()
                #db.session.commit()
            flash('Part updated successfully', 'success')
            return redirect(url_for('parts.list_parts'))
        except Exception as e:
            # Debug: Print the exception
            print(f"Error: {str(e)}")
            db.session.rollback()
            flash(f'Error updating part: {str(e)}', 'error')
            return redirect(url_for('parts.edit_part', part_id=part_id))

    
    categories = Category.query.all()
    suppliers = Supplier.query.all()
    warehouses = Warehouse.query.all()
    return render_template('parts/edit.html',
                         part=part,
                         categories=categories,
                         suppliers=suppliers,
                         warehouses=warehouses,
                         is_staff=is_staff)

@parts.route('/parts/<int:part_id>/bincard')
@login_required
def view_bincard(part_id):
    part = Part.query.get_or_404(part_id)
    entries = BinCard.query.filter_by(part_id=part_id).order_by(BinCard.date.desc()).all()
    
    return render_template('parts/bincard.html', 
                         part=part, 
                         entries=entries)

@parts.route('/api/check-part-number')
@login_required
def check_part_number():
    part_number = request.args.get('part_number')
    exists = Part.query.filter_by(part_number=part_number).first() is not None
    return jsonify({'exists': exists})

@parts.route('/parts/<int:part_id>/delete', methods=['POST'])
@login_required
def delete_part(part_id):
    part = Part.query.get_or_404(part_id)
    
    try:
        # Delete the part's image file if it exists
        if part.image_url:
            image_path = os.path.join(current_app.root_path, 'static', part.image_url.lstrip('/static/'))
            if os.path.exists(image_path):
                os.remove(image_path)
        
        # Delete associated bin card entries
        BinCard.query.filter_by(part_id=part_id).delete()
        
        # Delete the part
        db.session.delete(part)
        db.session.commit()
        flash('Part deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting part: {str(e)}', 'error')
        
    return redirect(url_for('parts.list_parts'))

@parts.route('/parts/<int:part_id>/detail')
@login_required
def view_part(part_id):
    part = Part.query.get_or_404(part_id)
    is_staff = is_user_staff()
    print(f"User {current_user.username} is staff: {is_staff}")  # Debug line
    
    return render_template('parts/detail.html',
                         part=part,
                         is_staff=is_staff)
@parts.route('/parts/export')
@login_required
@role_required('admin', 'manager')
def export_parts():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Parts List"

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        headers = [
            'Part Number',
            'Name',
            'Location',
            'Code',
            'Substitute Part Number',
            'Stock Level',
            'Min Stock',
            'Min Price',
            'Max Price',
            'Cost Price',
            'Description',
            'Unit',
            'Category',
            'Supplier',
            'Weight',
            'Height',
            'Length',
            'Width',
            'Color',
            'Warranty Period',
            'Barcode',
            'Created At',
            'Updated At'
        ]
        ws.append(headers)

        # Apply header styling
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Get all parts with their relationships
        parts = Part.query.join(Part.category).join(Part.supplier).all()

        # Write data rows
        for part in parts:
            row = [
                part.part_number,
                part.name,
                part.location,
                part.code,
                part.substitute_part_number,
                part.stock_level,
                part.min_stock,
                part.min_price,
                part.max_price,
                part.cost_price,
                part.cost_price_dirham,
                part.description,
                part.unit,
                part.category.name if part.category else '',
                part.supplier.name if part.supplier else '',
                part.weight,
                part.height,
                part.length,
                part.width,
                part.color,
                part.warranty_period,
                part.barcode,
                part.created_at.strftime('%Y-%m-%d %H:%M:%S') if part.created_at else '',
                part.updated_at.strftime('%Y-%m-%d %H:%M:%S') if part.updated_at else ''
            ]
            ws.append(row)

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Apply borders to all cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center')

        # Create a BytesIO object to save the workbook to
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'parts_list_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        flash(f'Error exporting parts: {str(e)}', 'error')
        return redirect(url_for('parts.list_parts'))

@parts.route('/api/part-names')
def part_names():
    q = request.args.get('q', '')
    results = PartName.query.filter(PartName.name.ilike(f'%{q}%')).limit(10).all()
    return jsonify([p.name for p in results])


@parts.route('/api/parts/search')
def search_parts():
    q = request.args.get('q', '')
    results = []
    if q:
        # Query your DB for matching parts (by part_number, name, etc)
        parts = Part.query.filter(
            (Part.part_number.ilike(f'%{q}%')) |
            (Part.name.ilike(f'%{q}%')) |
            (Part.code.ilike(f'%{q}%')) |
            (Part.substitute_part_number.ilike(f'%{q}%'))
        ).limit(30).all()
        for part in parts:
            results.append({
                'id': part.id,
                'part_number': part.part_number,
                'name': part.name,
                'code': part.code,
                'substitute_part_number': part.substitute_part_number
            })
    return jsonify({'results': results})